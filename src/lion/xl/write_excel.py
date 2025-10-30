from pathlib import Path
from lion.logger.exception_logger import log_exception
from openpyxl import load_workbook, Workbook, styles as openpyxl_styles
from openpyxl.utils.dataframe import dataframe_to_rows
from os import path
from lion.utils.popup_notifier import show_popup, show_error
from lion.utils.kill_file import kill_file
from lion.config.paths import LION_FILES_PATH
from shutil import copyfile
from pandas import DataFrame
import tempfile
from typing import Optional
from lion.utils.storage_manager import STORAGE_MANAGER, LionStorageManager


def save_wb(wb, xlpath):

    try:
        wb.save(xlpath)
    except Exception as er_msg:
        err = log_exception(popup=False)
        if 'permission denied' in err.lower():
            show_error(
                f"The file {xlpath} seems to be opened by another user or application." +
                " Please clsoe the file and Press OK to continue!")

            wb.save(xlpath)

        else:
            show_error(
                f'The file {xlpath} was not saved successfully!\n{str(er_msg)}')


def copy_file(xlpath=''):

    if str(xlpath).lower().endswith('movements.xlsm'):
        while True:
            try:
                copyfile(path.join(LION_FILES_PATH,
                         'movements.xlsm'), xlpath)
                break
            except Exception as er_msg:
                err = log_exception(popup=False)
                if 'permission denied' in err.lower():
                    show_error(
                        f"The file {xlpath} seems to be opened by another user or application." +
                        " Please clsoe the file and Press OK to continue!")

                    copyfile(path.join(LION_FILES_PATH,
                                       'movements.xlsm'), xlpath)

                else:
                    show_error(
                        f'The file {xlpath} was not saved successfully!\n{str(er_msg)}')


def create_blank_wb(xlpath):

    try:
        kill_file(xlpath)
    except Exception:
        pass

    if str(xlpath).lower().endswith('movements.xlsm'):
        copy_file(xlpath)
    else:
        wb = Workbook()
        save_wb(wb=wb, xlpath=xlpath)


def write_excel(df,
                xlpath='movements.xlsm',
                sheetname='NoName',
                keep=False,
                header=True,
                echo=True,
                storage: LionStorageManager | None = None):
    """
    Writes a dataframe to an Excel file locally or uploads to Azure Blob via LionStorageManager.
    
    """

    try:

        storage = storage or STORAGE_MANAGER

        is_blob_target = storage is not None and storage.in_azure
        local_tmp_path = None

        # If writing to blob, create a temporary local path first
        if is_blob_target:
            tmpdir = tempfile.gettempdir()
            local_tmp_path = Path(tmpdir) / Path(xlpath).name
            local_target = local_tmp_path
        else:
            local_target = Path(xlpath)

        if not keep:
            create_blank_wb(local_target)

        if str(local_target).lower().endswith(".xlsm"):
            wb = load_workbook(local_target, read_only=False, keep_vba=True)
        else:
            if not path.exists(local_target):
                create_blank_wb(local_target)
            wb = load_workbook(local_target, read_only=False)

        # write dataframe
        if sheetname:
            if sheetname in wb.sheetnames:
                ws = wb[sheetname]
                for row in ws.iter_rows(min_row=1 + int(not header)):
                    for cell in row:
                        cell.value = None
            elif "Sheet" in wb.sheetnames:
                s_sheet = wb["Sheet"]
                s_sheet.title = sheetname
                ws = wb[sheetname]
            else:
                ws = wb.create_sheet(sheetname)

            rows = dataframe_to_rows(df, index=False, header=header)
            for r_idx, row in enumerate(rows, 1 + int(not header)):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            hdr_fill = openpyxl_styles.PatternFill(
                patternType="solid", start_color="4D148C"
            )

            for col in ws.columns:
                col = [cell for cell in col]
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                for cell in col:
                    cell.font = openpyxl_styles.Font(name="FedEx Sans Cd", size=10)
                ws.column_dimensions[col[0].column_letter].width = max_len + 5

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            for cell in ws["1:1"]:
                cell.fill = hdr_fill
                cell.font = openpyxl_styles.Font(
                    color="FFFFFF", bold=True, name="FedEx Sans Cd", size=10
                )

            save_wb(wb=wb, xlpath=local_target)

        # if Azure, upload to blob
        if is_blob_target:
            blob_path = xlpath if isinstance(xlpath, str) else str(xlpath)
            storage.upload_file(local_tmp_path, container_name='logs', *blob_path.split("/"))

    except Exception:
        show_error(f"Creating {path.basename(str(xlpath))} failed! {log_exception()}")
        return False
    finally:
        if is_blob_target and local_tmp_path:
            local_tmp_path.unlink(missing_ok=True)

    if echo:
        show_popup(f"Data has been successfully exported to {xlpath}!")

    return True

def xlwriter(df, xlpath='movements.xlsm',
                sheetname='NoName', keep=False,
                header=True, echo=True):
    
    return write_excel(df=df, xlpath=xlpath,
                sheetname=sheetname, keep=keep,
                header=header, echo=echo)


if __name__ == '__main__':
    dict_ = {'key 1': 'value 1', 'key 2': 'value 2', 'key 3': 'value 3'}
    write_excel(df=DataFrame([dict_]))
