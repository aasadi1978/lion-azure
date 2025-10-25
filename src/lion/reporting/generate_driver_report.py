from collections import defaultdict
import logging
from lion.bootstrap.constants import LION_STRG_CONTAINER_DRIVER_REPORT
from lion.config.paths import LION_CONSOLIDATED_REPORT_PATH, LION_FONTS_PATH
from lion.shift_data.shift_data import UI_SHIFT_DATA
from lion.utils import utcnow
from lion.utils.km2mile import km2mile
from lion.utils.minutes2hhmm import minutes2hhmm
from lion.utils.is_null import is_null
from lion.utils.safe_copy import secure_copy
from lion.utils.kill_file import kill_file
from lion.utils.df2csv import export_dataframe_as_csv
from lion.logger.exception_logger import log_exception
from openpyxl import load_workbook, Workbook
import openpyxl
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import Image, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import inch
from pandas import DataFrame
from lion.orm.traffic_type import TrafficType
from lion.utils.concat import concat
from lion.orm.location import Location
from lion.orm.vehicle_type import VehicleType
from lion.orm.drivers_info import DriversInfo
from lion.config.paths import LION_LOCAL_DRIVER_REPORT_PATH
from lion.create_flask_app.create_app import LION_FLASK_APP
from os import path as os_path, makedirs
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from lion.utils.dict2class import Dict2Class
from lion.utils.get_week_num import get_week_num
from lion.config.paths import LION_IMAGES_PATH
from lion.logger.trigger_async_log_upload import trigger_async_log_upload

LION_STATION_REPORT_17Inch_Spotfire = 'https://euremars-lvl1-tss.emea.fedex.com:8001/spotfire/wp/OpenAnalysis?file=ec200e26-aa13-41ce-a2f0-6c291b5fd9ce'


class DriverReport():

    _instance = None
    def __new__(cls):

        if cls._instance is None:
            cls._instance = super().__new__(cls)
        
        return cls._instance

    def __init__(self):
        pass

    def initialize(self):

        try:
            self.__df_driver_report_base = DataFrame()
            self.__active_user_name = LION_FLASK_APP.config['LION_USER_FULL_NAME']
            self.__dct_footprint = Location.to_dict()
            self.__my_hdr_fill = openpyxl.styles.fills.PatternFill(
                patternType='solid', start_color='FF6600')

            self.__my_hdr_fill2 = openpyxl.styles.fills.PatternFill(
                patternType='solid', start_color='4D148C')

            self.__my_ftnote_fill = openpyxl.styles.fills.PatternFill(
                patternType='solid', start_color='999999')

            self.__str_today = datetime.now().strftime("%a %b %d %Y")
            self.__year = datetime.now().strftime('%Y')

            self.__right_click_id = False
            self.__page_export = False

            self.__dct_drivers = DriversInfo.to_dict()
            self.__dct_traffic_type_short_name = TrafficType.dct_short_names()

        except Exception as e:
            logging.error(f"Failed to init DriverReport! {e}")
            return
        
    @classmethod
    def get_instance(cls):
        return cls()

    @property
    def dict_all_movements_base(self):
        return self.__dict_all_movements_base

    @dict_all_movements_base.setter
    def dict_all_movements_base(self, x):
        self.__dict_all_movements_base = x

    @property
    def page_export(self):
        return self.__page_export

    @page_export.setter
    def page_export(self, x):
        self.__page_export = x

    @property
    def right_click_id(self):
        return self.__right_click_id

    @right_click_id.setter
    def right_click_id(self, x):
        self.__right_click_id = x

    @property
    def dump_directory(self):
        return self.__dump_directory

    @dump_directory.setter
    def dump_directory(self, x):
        self.__dump_directory = x

    def to_blob_storage(self, *blob_parts: str):
        
        try:

            utctimenow = utcnow()
            blob_parts: list[str] = list(blob_parts)
            blob_parts.append(utctimenow)

            trigger_async_log_upload(
                src_path=LION_STRG_CONTAINER_DRIVER_REPORT,
                container_name=LION_STRG_CONTAINER_DRIVER_REPORT,
                *blob_parts,
            )
        
        except Exception:
            log_exception(popup=False, remarks="Failed to trigger async upload of driver report to blob storage.")

    def __get_stn_pdf_path(self, loc, day, driver):
        """
        There can be only one master plan to generate driver report.
        To generate reports for other master plans, they have to be
        saved as scenario

        """

        if self.__right_click_id or self.__page_export:

            __filename = '%s.%s.pdf' % (driver, day)
            __stn_pdf_path = os_path.join(self.__dump_directory, __filename)
            __loc_report_dir = f"{self.__dump_directory}"

            kill_file(__stn_pdf_path)

            return __stn_pdf_path, __filename, __loc_report_dir

        __filename = '%s.pdf' % (driver)
        wknum = get_week_num()

        __loc_report_dir = os_path.join(
            self.__dump_directory, r'Driver plan %s' % (wknum))

        __loc_report_dir = os_path.join(
            __loc_report_dir, r'%s' % (loc))

        __loc_report_dir = os_path.join(
            __loc_report_dir, r'%s' % (day))

        makedirs(__loc_report_dir, exist_ok=True)

        __stn_pdf_path = os_path.join(__loc_report_dir, __filename)

        kill_file(__stn_pdf_path)

        return __stn_pdf_path, __filename, __loc_report_dir

    def __get_stn_wb_path(self, loc, day, driver):
        """
        There can be only one amster plan to generate driver report.
        To generate reports for other master plans, they have to be
        saved as scenario
        """

        wknum = get_week_num()
        __loc_report_dir = os_path.join(
            self.__dump_directory, r'Driver plan %s' % (wknum))

        __loc_report_dir = os_path.join(
            __loc_report_dir, r'%s' % (loc))

        __loc_report_dir = os_path.join(
            __loc_report_dir, r'%s' % (day))

        makedirs(__loc_report_dir, exist_ok=True)

        __filename = '%s.xlsx' % (driver)

        __stn_wb_path = os_path.join(__loc_report_dir, __filename)

        kill_file(__stn_wb_path)

        wb = Workbook()
        try:
            wb.save(__stn_wb_path)
        except Exception:

            __now = datetime.now().strftime('%Y.%m.%d %H%M')
            __filename = '%s %s.%s.xlsx' % (driver, __now, day)
            __stn_wb_path = os_path.join(__loc_report_dir, __filename)
            wb.save(__stn_wb_path)

        return __stn_wb_path

    def __create_pdf_report(self, dataframe, remarks, pdf_path, title):

        try:

            pdfmetrics.registerFont(
                TTFont('FedEx Sans', LION_FONTS_PATH / 'fdxfonts' / 'FedExSans_Rg.ttf'))

            custom_style = ParagraphStyle(
                'CustomStyle',
                fontSize=9,
                fontName='FedEx Sans'
            )

            doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()

            elements.append(Spacer(1, 10))

            # Title Style
            title_style = styles['Heading2']
            title_style.alignment = TA_CENTER
            title_style.fontName = 'FedEx Sans'

            # Add Title
            elements.append(Paragraph(title, title_style))

            _columns = dataframe.columns.tolist()

            # Convert DataFrame to list of lists
            data = [_columns] + dataframe.values.tolist()

            # Create Table
            # table = Table(data)
            # _n_cols = len(_columns)
            col_widths = [1.5 * inch] * 3
            col_widths.extend([0.7 * inch] * 2)
            col_widths.append(0.75 * inch)
            col_widths.append(0.5 * inch)
            col_widths.append(1.05 * inch)
            col_widths.append(1.05 * inch)

            table = Table(data, colWidths=col_widths)

            # Table Style
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), '#4D148C'),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                # ('VALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'FedEx Sans'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, '#999999')
            ])

            table.setStyle(style)
            elements.append(table)
            elements.append(Spacer(1, 20))

            # Add Remarks
            remarks_style = styles['BodyText']
            remarks_style.alignment = TA_LEFT
            elements.append(Paragraph("Remarks: " + remarks, custom_style))

            custom_style2 = ParagraphStyle(
                'CustomStyle',
                fontSize=6,
                fontName='FedEx Sans'
            )

            elements.append(Spacer(1, 10))

            _author = f"© Copyright {self.__year} Road Network Operations UK, FedEx Express. Created by {
                self.__active_user_name} on {self.__str_today}. All rights reserved"

            par_author = Paragraph(_author, custom_style2)
            par_author.alignment = TA_CENTER
            elements.append(par_author)

            elements.append(Spacer(1, 10))

            # Add Images (replace with your image paths)
            image_path = os_path.join(
                LION_IMAGES_PATH, 'fdx.png')

            if os_path.exists(image_path):

                img = Image(image_path,  width=43, height=15)
                img.hAlign = 'RIGHT'
                elements.append(img)

            # image_path = os_path.join(
            #     LION_IMAGES_PATH, '2019_Rome_08526_Banner-601x259-6d194da.png')

            # if os_path.exists(image_path):
            #     img = Image(image_path,  width=650, height=150)
            #     img.hAlign = 'CENTER'
            #     elements.append(img)

            # Build PDF
            doc.build(elements)

        except Exception:
            return f'Driver report pdf generation failed: {log_exception(False)}'

        return ''

    def __report_xlsx_or_pdf(self, shift_id,
                             dataframe,
                             day,
                             dep_day,
                             remarks):

        try:

            __loc = self.__dct_drivers[shift_id]['controlled by']
            __vhcl = self.__dct_drivers[shift_id]['vehicle']
            __vehiclename = VehicleType.get_vehicle_name(__vhcl)
            driver = self.__dct_drivers[shift_id]['shiftname']

            __title = '%s (%s) on %s' % (
                driver, self.__dct_drivers[shift_id]['operator'], dep_day)

            pr_pdf_path, _, _ = self.__get_stn_pdf_path(
                loc=__loc, day=day, driver=driver)

            _try_excel = self.__create_pdf_report(
                dataframe=dataframe.copy(), pdf_path=pr_pdf_path,
                remarks=remarks, title='%s (%s)' % (__title, __vehiclename)) != ''

            if _try_excel:

                n_rows = dataframe.shape[0]
                __first_row = 7
                __font = Font(name='FedEx Sans', sz=9)

                __stn_wb_path = self.__get_stn_wb_path(
                    loc=__loc, day=day, driver=driver)

                wb = load_workbook(__stn_wb_path, read_only=False)

                if 'Sheet' in wb.sheetnames:
                    s_sheet = wb['Sheet']
                    s_sheet.title = driver
                    ws = wb[driver]

                else:
                    ws = wb.create_sheet(driver)

                rows = dataframe_to_rows(dataframe, index=False, header=True)
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx + __first_row,
                                column=c_idx, value=value)

                ws["A7"].fill = self.__my_hdr_fill
                ws["B7"].fill = self.__my_hdr_fill
                ws["C7"].fill = self.__my_hdr_fill
                ws["D7"].fill = self.__my_hdr_fill
                ws["E7"].fill = self.__my_hdr_fill
                ws["F7"].fill = self.__my_hdr_fill
                ws["G7"].fill = self.__my_hdr_fill
                ws["H7"].fill = self.__my_hdr_fill
                ws["I7"].fill = self.__my_hdr_fill

                ws["A8"].fill = self.__my_hdr_fill2
                ws["A8"].font = Font(color='00FFFFFF', bold=True)

                ws["B8"].fill = self.__my_hdr_fill2
                ws["B8"].font = Font(color='00FFFFFF', bold=True)

                ws["C8"].fill = self.__my_hdr_fill2
                ws["C8"].font = Font(color='00FFFFFF', bold=True)

                ws["D8"].fill = self.__my_hdr_fill2
                ws["D8"].font = Font(color='00FFFFFF', bold=True)

                ws["E8"].fill = self.__my_hdr_fill2
                ws["E8"].font = Font(color='00FFFFFF', bold=True)

                ws["F8"].fill = self.__my_hdr_fill2
                ws["F8"].font = Font(color='00FFFFFF', bold=True)

                ws["G8"].fill = self.__my_hdr_fill2
                ws["G8"].font = Font(color='00FFFFFF', bold=True)

                ws["H8"].fill = self.__my_hdr_fill2
                ws["H8"].font = Font(color='00FFFFFF', bold=True)

                ws["I8"].fill = self.__my_hdr_fill2
                ws["I8"].font = Font(color='00FFFFFF', bold=True)

                ws.sheet_view.showGridLines = False

                __title_cell = 'C%d' % (__first_row)
                ws[__title_cell] = '%s (%s)' % (__title, __vehiclename)
                ws[__title_cell].font = Font(
                    name='FedEx Sans Cd Medium', sz=14, bold=True)

                for cell in ws["%d:%d" % (__first_row, __first_row)]:
                    cell.font = Font(bold=True,  name='FedEx Sans', )

                __remarks_row = n_rows + __first_row + 4

                ws['A%d' % (__remarks_row)] = 'Remarks:'
                ws['A%d' % (__remarks_row)].font = Font(bold=True)

                for j in range(__first_row+2, __remarks_row+1):
                    for cell in ws["%d:%d" % (j, j)]:
                        cell.font = __font

                __remars_list = remarks.split(' ')
                __r = __remarks_row + 1
                while __remars_list:

                    ws['A%d' % (__r)] = ' '.join(__remars_list[: 30])
                    __remars_list = __remars_list[30:]
                    ws['A%d' % (__r)].font = __font
                    __r += 1

                _author = f"Created on {
                    self.__str_today} by {self.__active_user_name}"
                _author = f"{_author} © {
                    self.__year} FedEx Express. All rights reserved"

                __fotnt_row = __r + 2
                ws['A%d' % (__fotnt_row)] = _author

                for cell in ws["%d:%d" % (__fotnt_row, __fotnt_row)]:
                    cell.font = Font(size=8, name='FedEx Sans')
                    cell.fill = self.__my_ftnote_fill

                img = openpyxl.drawing.image.Image(
                    os_path.join(LION_IMAGES_PATH, 'fdx.png'))

                img.anchor = 'A1'
                ws.add_image(img)

                img2 = openpyxl.drawing.image.Image(
                    os_path.join(LION_IMAGES_PATH, 'noPrnt.png'))
                img2.anchor = 'H1'
                ws.add_image(img2)

                img3 = openpyxl.drawing.image.Image(
                    os_path.join(LION_IMAGES_PATH, 'break2.png'))

                img3.height = 58
                img3.width = 95
                img3.anchor = 'I5'
                ws.add_image(img3)

                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 12

                ws.page_setup.orientation = 'landscape'
                ws.page_setup.paperSize = '9'  # A4
                ws.page_setup.fitToPage = True

                ws.protection.password = 'LionUK@2023'

            else:
                return ''

        except Exception:
            return 'Driver report xlsx generation failed. %s' % (log_exception(False))

        wb.save(__stn_wb_path)
        return ''

    def __set_base_data(self):

        try:
            self.__dct_drivers = DriversInfo.to_dict()
            self._dct_shift_running_on_weekday = DriversInfo.get_dct_shifts_per_weekday()

            self.__dct_optimal_drivers_base = UI_SHIFT_DATA.optimal_drivers
            self.__dict_all_movements_base = UI_SHIFT_DATA.dict_all_movements

            self.__dct_optimal_drivers_base.pop(1, None)
            self.__dct_optimal_drivers_base.pop(2, None)

            self.__df_driver_report_base = DataFrame()

        except Exception:
            log_exception(
                popup=False, remarks='Setting base data failed for reporting!')

    def gen_report_base(self, shift_ids=[]):

        self.__set_base_data()

        __dct_shift_breaks = {}
        __unprocessed_drivers = []
        self.__dct_footprint = Location.to_dict()

        try:

            if not shift_ids:
                drivers = list(self.__dct_optimal_drivers_base)
            else:
                drivers = [
                    d for d in shift_ids if d in self.__dct_optimal_drivers_base]

            __dct_optimal_drivers_base = {
                d: v for d, v in self.__dct_optimal_drivers_base.items() if d in drivers}

            _dct_optimal_drivers = secure_copy(
                __dct_optimal_drivers_base)

            _dict_all_movements = secure_copy(
                self.__dict_all_movements_base)

        except Exception:

            _dct_optimal_drivers = {}
            _dict_all_movements = {}

            log_exception(
                popup=True, remarks='Pre-processing data failed')

        if _dct_optimal_drivers:

            __df_movements = DataFrame.from_dict(
                _dict_all_movements, orient='index')

            __df_movements = __df_movements.loc[:, [
                'MovementID', 'From', 'To', 'DepDateTime',
                'ArrDateTime', 'DrivingTime', 'Dist']].copy()

            __df_driver_report = DataFrame(columns=[
                'Driver',
                'shift_id',
                'TourSequence',
                'MovementID',
                'vehicle',
                'From',
                'To',
                'DepDateTime',
                'ArrDateTime',
                'DrivingTime',
                'Dist',
                'Mile',
                'tu_loc',
                'total_dist_cost'
            ])

            __dct_shift_note = defaultdict()
            __dct_shift_remarks = defaultdict()

            for driver in set(_dct_optimal_drivers):

                __movs = _dct_optimal_drivers[driver].get(
                    'list_movements', [])

                try:

                    if not __movs:
                        raise ValueError(
                            f'The shift {driver} is empty!')

                    __dct_running_data = _dct_optimal_drivers[driver]['dct_running_tour_data']

                    for __m1 in __dct_running_data:
                        __brk = __dct_running_data[__m1]['Break']
                        if __brk:
                            loc = __dct_running_data[__m1]['loc']
                            __dct_shift_breaks[__m1] = '@%s: %d min' % (
                                loc, int(__brk))

                    __movs = _dct_optimal_drivers[driver]['list_movements']

                    __dct_shift_note[driver] = _dct_optimal_drivers[driver].get(
                        'shift_note', '')

                    __dct_shift_remarks[driver] = _dct_optimal_drivers[driver].get(
                        'remark', '')

                    __dct_movs_idx = {
                        m: idx + 1 for idx, m in enumerate(__movs)}

                    __df = __df_movements[
                        __df_movements.MovementID.isin(__movs)].copy()

                    __df['TourSequence'] = __df.MovementID.apply(
                        lambda x: __dct_movs_idx[x])

                    __df['vehicle'] = VehicleType.get_vehicle_name(
                        _dct_optimal_drivers[driver]['vehicle'])

                    __df['Driver'] = self.__dct_drivers[driver]['shiftname']
                    __df['shift_id'] = driver

                    __df['Mile'] = __df.Dist.apply(
                        lambda x: km2mile(x))

                    __df['tu_loc'] = __df.MovementID.apply(
                        lambda x: _dict_all_movements[x].get('tu', ''))

                    __df['DrivingTime'] = __df.DrivingTime.apply(
                        lambda x: "" + minutes2hhmm(x))

                    __df['Break'] = __df.MovementID.apply(
                        lambda x: __dct_shift_breaks.get(x, ''))

                    __df_driver_report = concat(
                        [__df_driver_report, __df])

                except Exception:
                    log_exception(popup=False)
                    __unprocessed_drivers.append(str(driver))

            try:
                __df_driver_report['Traffic Type'] = __df_driver_report.MovementID.apply(
                    lambda x: _dict_all_movements[x].get('TrafficType', ''))

                __df_driver_report['Traffic Type'] = __df_driver_report['Traffic Type'].apply(
                    lambda x: self.__dct_traffic_type_short_name.get(x, x))

                __df_driver_report.sort_values(
                    by=['Driver', 'TourSequence'], ascending=True, inplace=True)

                __df_driver_report = __df_driver_report.loc[:, [
                    'MovementID',
                    'Driver',
                    'shift_id',
                    'From',
                    'To',
                    'tu_loc',
                    'DepDateTime',
                    'ArrDateTime',
                    'DrivingTime',
                    'Mile',
                    'Traffic Type',
                    'Break'
                ]].copy()

                set_Locs = set(__df_driver_report.From.tolist())
                set_Locs.update(__df_driver_report.To)

                _missing_locs = [
                    x for x in set_Locs if x not in self.__dct_footprint]

                if _missing_locs:
                    raise ValueError(
                        f"{len(_missing_locs)} unknown locations: {'|'.join(list(_missing_locs[:10]))} ... !")

                __df_driver_report['From'] = __df_driver_report.From.apply(
                    lambda x: '%s/%s' % (x, self.__dct_footprint[x]['location_name']))

                __df_driver_report['To'] = __df_driver_report.To.apply(
                    lambda x: '%s/%s' % (x, self.__dct_footprint[x]['location_name']))

                __df_driver_report['tu_loc'] = __df_driver_report.tu_loc.apply(
                    lambda x: '' if is_null(x) else '%s/%s' % (x, self.__dct_footprint.get(x, {}).get('location_name', x)))

                __df_driver_report['DepDateTime'] = __df_driver_report.DepDateTime.apply(
                    lambda x: x.strftime('%a %H:%M')
                )

                __df_driver_report['ArrDateTime'] = __df_driver_report.ArrDateTime.apply(
                    lambda x: x.strftime('%a %H:%M')
                )

                for strng in ['ROAD HUB', 'AIR HUB']:

                    __df_driver_report['From'] = __df_driver_report.From.apply(
                        lambda x: x.replace(strng, '').strip())

                    __df_driver_report['To'] = __df_driver_report.To.apply(
                        lambda x: x.replace(strng, '').strip())

                    __df_driver_report['tu_loc'] = __df_driver_report.tu_loc.apply(
                        lambda x: x.replace(strng, '').strip())

                __df_driver_report.rename(
                    columns={
                        'DrivingTime': 'Driving Time',
                        'From': 'Start Point',
                        'To': 'End Point',
                        'DepDateTime': 'Start Time',
                        'ArrDateTime': 'End Time',
                        'Mile': 'Distance (miles)',
                        'tu_loc': 'TU Destination'
                    },
                    inplace=True
                )

                __df_driver_report['remarks'] = __df_driver_report.shift_id.apply(
                    lambda x: __dct_shift_remarks[x])

                __df_driver_report['Driving Time'] = __df_driver_report['Driving Time'].apply(
                    lambda x: "%s" % (x))

                __df_driver_report['driver_loc'] = __df_driver_report.shift_id.apply(
                    lambda x: self.__dct_drivers[x]['start position'])

                __df_driver_report['lat'] = __df_driver_report.driver_loc.apply(
                    lambda x: self.__dct_footprint.get(x, {}).get('latitude', 0))

                __df_driver_report['lon'] = __df_driver_report.driver_loc.apply(
                    lambda x: self.__dct_footprint.get(x, {}).get('longitude', 0))

                __df_driver_report['driver_loc_name'] = __df_driver_report.driver_loc.apply(
                    lambda x: self.__dct_footprint.get(x, {}).get('location_name', x))

                __df_driver_report['notes'] = __df_driver_report.shift_id.apply(
                    lambda x: __dct_shift_note[x])

                __locs = __df_driver_report.driver_loc.tolist()
                __df_driver_report['nShifts'] = __df_driver_report.driver_loc.apply(
                    lambda x: __locs.count(x))

                __df_driver_report['vehicle'] = __df_driver_report.shift_id.apply(
                    lambda x: VehicleType.get_vehicle_name(_dct_optimal_drivers[x]['vehicle']))

                __df_driver_report['operator'] = __df_driver_report.shift_id.apply(
                    lambda x: self.__dct_drivers[x]['operator'])

                __df_driver_report['last_update'] = datetime.now()

            except Exception:
                log_exception(popup=False)

            self.__df_driver_report_base = __df_driver_report.copy()

    def gen_report(self, drivers=[], return_driver_report=False, wkday='Mon', no_pdf=False):

        __dct_output = {}
        __status_msg = ''

        publish = self.__dump_directory != LION_LOCAL_DRIVER_REPORT_PATH

        __unprocessed_drivers = []
        self.__dct_footprint = Location.to_dict()

        try:

            if self.__df_driver_report_base.empty:
                self.gen_report_base()

            if self.__df_driver_report_base.empty:
                raise ValueError('gen_report_base was not created!')

            if not drivers:
                drivers = list(self.__dct_optimal_drivers_base)

            drivers = [
                d for d in drivers if not self.__dct_optimal_drivers_base[d].is_blank]

            drivers = [d for d in drivers if DriversInfo.shift_id_runs_on_weekday(
                shift_id=d, weekday=wkday) and not self.__dct_optimal_drivers_base[d].is_blank]

            if not drivers:

                if return_driver_report:
                    return DataFrame()

                return Dict2Class({'status_msg': ''})

            __df_driver_report = self.__df_driver_report_base[
                self.__df_driver_report_base.shift_id.isin(drivers)].copy()

            if __df_driver_report.empty:
                raise ValueError(
                    'No report available for the specified shifts!')

            __df_driver_report['Start Time'] = __df_driver_report.MovementID.apply(
                lambda x: self.__dict_all_movements_base[x].get_modified_dep_datetimes(wkday=wkday).strftime('%a %H:%M'))

            __df_driver_report['End Time'] = __df_driver_report.MovementID.apply(
                lambda x: self.__dict_all_movements_base[x].get_modified_arr_datetimes(wkday=wkday).strftime('%a %H:%M'))

            __df_driver_report['dep_day'] = __df_driver_report.shift_id.apply(
                lambda x: self.__dct_optimal_drivers_base[x].modify_dep_day(wkday=wkday).strftime('%A'))

            __df_driver_report['weekday'] = wkday

            report_columns = ['MovementID', 'Driver', 'shift_id', 'Start Point', 'End Point', 'TU Destination',
                              'Start Time', 'End Time', 'Driving Time', 'Distance (miles)', 'Traffic Type',
                              'Break', 'remarks', 'dep_day']

            __df_driver_report_pdf = __df_driver_report.loc[
                :, report_columns].copy()

            __df_driver_report_pdf.rename(
                columns=({'Distance (miles)': 'Miles'}), inplace=True)

            __df_driver_report_pdf['Start Point'] = __df_driver_report_pdf.apply(
                lambda x: x['Start Point'] if len(
                    x['Start Point']) <= 20 else x['Start Point'][:19] + '.', axis=1
            )

            __df_driver_report_pdf['End Point'] = __df_driver_report_pdf.apply(
                lambda x: x['End Point'] if len(
                    x['End Point']) <= 20 else x['End Point'][:19] + '.', axis=1
            )

            __df_driver_report_pdf['TU Destination'] = __df_driver_report_pdf.apply(
                lambda x: x['TU Destination'] if len(
                    x['TU Destination']) <= 20 else x['TU Destination'][:19] + '.', axis=1
            )

        except Exception:

            __df_driver_report = DataFrame()
            __df_driver_report_pdf = DataFrame()
            log_exception(
                popup=False, remarks='Pre-processing data failed')

        try:

            dump_in_day_dir = False

            if return_driver_report:

                try:

                    _df = __df_driver_report.copy()
                    _df['title'] = _df.apply(
                        lambda x: '%s (%s) on %s' % (
                            x['Driver'], self.__dct_drivers[x['shift_id']]['operator'], x['dep_day']), axis=1)

                    _df['last_update'] = datetime.now().strftime(
                        '%Y-%m-%d %H:%M')
                    _df.rename(columns={c: c.lower().replace(' ', '_')
                                        for c in _df.columns.tolist()}, inplace=True)

                    _df.rename(columns={
                        'tu_destination': 'tu_dest',
                        'distance_(miles)': 'dist_miles',
                        'break': 'break_info'
                    }, inplace=True)

                    return _df

                except Exception:
                    log_exception(popup=False)

                    return DataFrame()

            else:

                if publish:
                    export_dataframe_as_csv(dataframe=__df_driver_report.copy(), csv_file_path=LION_CONSOLIDATED_REPORT_PATH /f'df_driver_report_{wkday}.csv')

                else:
                    export_dataframe_as_csv(dataframe=__df_driver_report.copy(), csv_file_path=self.__dump_directory /f'df_driver_report_{wkday}.csv')

            if not no_pdf:

                __cntr: int = 0
                __status_msg = ''

                for _shiftid in drivers:

                    __cntr += 1
                    __df = __df_driver_report_pdf[
                        __df_driver_report_pdf.shift_id == _shiftid].copy()

                    depday = __df.dep_day.tolist()[0]
                    remarks = __df.remarks.tolist()[0]

                    __df.drop(
                        columns=['Driver', 'shift_id', 'dep_day',
                                 'remarks', 'MovementID'], axis=1, inplace=True)

                    __status = self.__report_xlsx_or_pdf(shift_id=_shiftid,
                                                         dataframe=__df,
                                                         dep_day=depday,
                                                         day=wkday,
                                                         remarks=remarks)

                    if __status != '':
                        __status_msg = __status_msg + __status + '; '

            if __unprocessed_drivers:
                __status_msg = __status_msg + 'No driver plan was generated for %d shifts on %s: %s' % (
                    len(__unprocessed_drivers), wkday, ';'.join(__unprocessed_drivers))

        except Exception:
            __status_msg = f"{__status_msg}\n{log_exception(
                popup=False, remarks=f'Generating {wkday} driver report failed!')}"

            if return_driver_report:
                return DataFrame()

        __dct_output['status_msg'] = __status_msg
        __dct_output['unprocessed_drivers'] = __unprocessed_drivers

        if __status_msg == '':
            __status_msg = 'Driver report finished successfully for %s.' % wkday

        return Dict2Class(__dct_output)

DRIVER_REPORT = DriverReport.get_instance()